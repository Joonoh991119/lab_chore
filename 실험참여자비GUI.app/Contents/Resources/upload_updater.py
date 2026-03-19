"""
업로드 양식 행 추가 자동화  —  CSNL lab_chore
사용법:
  python3 upload_updater.py --all          # 폴더 내 전체
  python3 upload_updater.py 파일.xlsx      # 개별 지정
  python3 upload_updater.py               # 파일 선택 다이얼로그
의존: pip install openpyxl
"""
import sys, os, shutil, zipfile, io, unicodedata
from copy import copy
import openpyxl

WORK_DIR = os.environ.get("LAB_CHORE_DIR", os.path.dirname(os.path.abspath(__file__)))
UPLOAD_TEMPLATE = os.path.join(WORK_DIR, "template_일회성경비지급자 업로드양식.xlsx")
UPLOAD_BACKUP   = os.path.join(WORK_DIR,
    "template_일회성경비지급자 업로드양식.xlsx.sb-ce0a4f46-hQO6zg", "8C283410.MACTF")
OUTPUT_UPLOAD   = os.path.join(WORK_DIR, "일회성경비지급자_업로드양식_작성.xlsx")

DEFAULT_NATIONALITY   = "대한민국"
DEFAULT_INCOME_TYPE   = "기타소득"
DEFAULT_INCOME_DETAIL = "강연료 등 필요경비 있는 기타소득"

def _valid_xlsx(p):
    try:
        with open(p,"rb") as f: zipfile.ZipFile(io.BytesIO(f.read()))
        return True
    except: return False

def load_wb():
    if os.path.exists(OUTPUT_UPLOAD) and _valid_xlsx(OUTPUT_UPLOAD):
        return openpyxl.load_workbook(OUTPUT_UPLOAD)
    if os.path.exists(UPLOAD_TEMPLATE) and _valid_xlsx(UPLOAD_TEMPLATE):
        shutil.copy2(UPLOAD_TEMPLATE, OUTPUT_UPLOAD)
    elif os.path.exists(UPLOAD_BACKUP):
        with open(UPLOAD_BACKUP,"rb") as f: data = f.read()
        with open(OUTPUT_UPLOAD,"wb") as f: f.write(data)
    else:
        raise FileNotFoundError(f"업로드 양식 템플릿을 찾을 수 없습니다.\n원본: {UPLOAD_TEMPLATE}\n백업: {UPLOAD_BACKUP}")
    print(f"[INFO] 출력 파일 생성: {OUTPUT_UPLOAD}")
    # 샘플 데이터 제거 후 저장
    wb_new = openpyxl.load_workbook(OUTPUT_UPLOAD)
    _init_fresh_output(wb_new["Sheet1"])
    wb_new.save(OUTPUT_UPLOAD)
    return openpyxl.load_workbook(OUTPUT_UPLOAD)

def read_info(path):
    wb = openpyxl.load_workbook(path, data_only=True); ws = wb.active
    name    = str(ws["B16"].value or "").strip()
    inst    = str(ws["D16"].value or "").strip()
    jid_raw = str(ws["E16"].value or "").strip()
    bank    = str(ws["G16"].value or "").strip()
    account = str(ws["I16"].value or "").strip().replace(" ","")
    holder  = str(ws["L16"].value or name).strip()
    amount  = ws["D19"].value
    jc = jid_raw.replace("-","").replace(" ","")
    jf, jb = (jc[:6], jc[6:13]) if len(jc)>=13 else \
             (jid_raw.split("-")[0], jid_raw.split("-")[1]) if "-" in jid_raw else (jc,"")
    try: amount = int(amount)
    except: amount = 0
    return dict(name=name,inst=inst,jid_front=jf,jid_back=jb,
                bank=bank,account=account,holder=holder,amount=amount)

def end_row(ws):
    for r in ws.iter_rows(min_col=1,max_col=1):
        if str(r[0].value).strip().upper()=="END": return r[0].row
    return ws.max_row+1

def _init_fresh_output(ws):
    """샘플 데이터(금동이·은동이 등) 제거, row 3 스타일 보존, row 4 = END."""
    er = end_row(ws)
    for col in range(1, 19):
        ws.cell(3, col).value = None     # 값만 제거, 스타일 유지
    rows_to_delete = er - 4
    if rows_to_delete > 0:
        ws.delete_rows(4, rows_to_delete)
    ws.cell(4, 1).value = "END"

def copy_row_style(ws, src_row: int, dst_row: int, max_col: int = 18):
    """src_row의 폰트·테두리·채우기·정렬·행높이를 dst_row에 복사."""
    src_dim = ws.row_dimensions[src_row]
    ws.row_dimensions[dst_row].height = src_dim.height
    for col in range(1, max_col + 1):
        src = ws.cell(src_row, col)
        dst = ws.cell(dst_row, col)
        if src.has_style:
            dst.font          = copy(src.font)
            dst.border        = copy(src.border)
            dst.fill          = copy(src.fill)
            dst.number_format = src.number_format
            dst.protection    = copy(src.protection)
            dst.alignment     = copy(src.alignment)

def next_seq(wb):
    ws=wb["Sheet1"]; er=end_row(ws)
    for r in range(er-1, 2, -1):
        try: return int(ws.cell(r,1).value)+1
        except (TypeError, ValueError): continue
    return 1

def append_row(wb, info, seq):
    ws  = wb["Sheet1"]
    r   = end_row(ws)
    ref = (r - 1) if r > 3 else 3   # 스타일 참조 행
    ws.insert_rows(r)
    copy_row_style(ws, ref, r)       # 스타일 먼저 복사
    vals = [seq, info["name"], info["inst"], info["jid_front"], info["jid_back"],
            f'=IF(H{r}="대한민국","N","Y")', "", DEFAULT_NATIONALITY,
            DEFAULT_INCOME_TYPE, DEFAULT_INCOME_DETAIL, info["amount"],
            info["account"], info["bank"], info["holder"], 0, 0, 0, 0]
    for i,v in enumerate(vals,1): ws.cell(r,i).value=v
    print(f"  → 행{r}: {info['name']} | {info['amount']:,}원 | {info['bank']} {info['account']}")

def process(files):
    wb = load_wb(); ws=wb["Sheet1"]
    existing={str(r[1]).strip() for r in ws.iter_rows(min_row=3,max_col=2,values_only=True)
              if r[0] and str(r[0]).upper()!="END" and r[1]}
    added=skipped=0
    for path in files:
        path=os.path.abspath(path); fname=os.path.basename(path)
        print(f"\n[처리] {fname}")
        if not os.path.exists(path): print("  [WARN] 없음"); skipped+=1; continue
        try: info=read_info(path)
        except Exception as e: print(f"  [ERROR] {e}"); skipped+=1; continue
        if not info["name"]: print("  [WARN] 이름 없음"); skipped+=1; continue
        if info["name"] in existing: print(f"  [SKIP] 이미 등록됨"); skipped+=1; continue
        append_row(wb, info, next_seq(wb)); existing.add(info["name"]); added+=1
    wb.save(OUTPUT_UPLOAD)
    print(f"\n{'─'*50}\n✅ {added}명 추가 / {skipped}건 건너뜀\n📄 {OUTPUT_UPLOAD}")

def main():
    if "--all" in sys.argv:
        prefix="실험참여자비 양식_"
        files=sorted([os.path.join(WORK_DIR,f) for f in os.listdir(WORK_DIR)
                      if f.endswith(".xlsx") and unicodedata.normalize("NFC",f).startswith(prefix)])
        if not files: print("[INFO] 처리할 파일 없음"); sys.exit(0)
        print(f"[INFO] {len(files)}개 처리")
        for f in files: print(f"  • {os.path.basename(f)}")
        process(files)
    elif len(sys.argv)>=2:
        files=[os.path.join(WORK_DIR,f) if not os.path.isabs(f) else f
               for f in sys.argv[1:] if not f.startswith("--")]
        process(files)
    else:
        import subprocess
        res=subprocess.run(["osascript","-e",
            f'set f to choose file with prompt "참가자 양식 파일 선택" of type {{"xlsx"}} '
            f'default location POSIX file "{WORK_DIR}" with multiple selections allowed'],
            capture_output=True,text=True)
        if res.returncode!=0: print("[INFO] 취소됨"); sys.exit(0)
        paths=[]
        for ap in res.stdout.strip().split(", "):
            pr=subprocess.run(["osascript","-e",f'POSIX path of ("{ap.strip()}")'],
                               capture_output=True,text=True)
            paths.append(pr.stdout.strip())
        process(paths)

if __name__=="__main__": main()
