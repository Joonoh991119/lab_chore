"""
실험참여자비 양식 자동 입력 GUI  —  CSNL lab_chore  v3.0
실행: .app 더블클릭  또는  streamlit run app.py
의존: pip install streamlit openpyxl streamlit-drawable-canvas Pillow
"""
import os, re, shutil, io, datetime, subprocess, unicodedata, zipfile
import openpyxl
from openpyxl.drawing.image import Image as XLImage
import streamlit as st
from PIL import Image

# ═══════════════════════════════════════════════════════════════════════════
#  공통 유틸
# ═══════════════════════════════════════════════════════════════════════════
BANK_LIST = [
    "국민은행","기업은행","신한은행","우리은행","하나은행",
    "농협은행","SC제일은행","씨티은행","카카오뱅크","토스뱅크",
    "케이뱅크","부산은행","대구은행","경남은행","광주은행",
    "전북은행","제주은행","산업은행","수협은행",
    "새마을금고","신협","우체국","저축은행","기타",
]
DEFAULT_NATIONALITY   = "대한민국"
DEFAULT_INCOME_TYPE   = "기타소득"
DEFAULT_INCOME_DETAIL = "강연료 등 필요경비 있는 기타소득"


def pick_folder_macos(default: str) -> str | None:
    """osascript 폴더 선택 다이얼로그 → POSIX path 반환, 취소시 None."""
    try:
        r1 = subprocess.run(
            ["osascript", "-e",
             f'set f to choose folder with prompt "템플릿 파일 폴더를 선택하세요"'
             f' default location POSIX file "{default}"'],
            capture_output=True, text=True, timeout=60,
        )
        if r1.returncode != 0:
            return None
        alias = r1.stdout.strip()
        r2 = subprocess.run(
            ["osascript", "-e", f'POSIX path of ("{alias}")'],
            capture_output=True, text=True, timeout=5,
        )
        path = r2.stdout.strip().rstrip("/")
        return path if path else None
    except Exception:
        return None


def _valid_xlsx(p: str) -> bool:
    try:
        with open(p, "rb") as f:
            zipfile.ZipFile(io.BytesIO(f.read()))
        return True
    except Exception:
        return False

def scan_participant_forms(work_dir: str) -> list[str]:
    """실험참여자비 양식_*.xlsx 파일 목록 반환 (NFD/NFC 양쪽 대응)."""
    prefix = "실험참여자비 양식_"
    return sorted([
        os.path.join(work_dir, f)
        for f in os.listdir(work_dir)
        if f.endswith(".xlsx")
        and unicodedata.normalize("NFC", f).startswith(prefix)
    ])


def read_participant_info(path: str) -> dict:
    wb = openpyxl.load_workbook(path, data_only=True)
    ws = wb.active
    name    = str(ws["B16"].value or "").strip()
    inst    = str(ws["D16"].value or "").strip()
    jid_raw = str(ws["E16"].value or "").strip()
    bank    = str(ws["G16"].value or "").strip()
    account = str(ws["I16"].value or "").strip().replace(" ", "")
    holder  = str(ws["L16"].value or name).strip()
    amount  = ws["D19"].value
    jc = jid_raw.replace("-", "").replace(" ", "")
    if len(jc) >= 13:
        jf, jb = jc[:6], jc[6:13]
    elif "-" in jid_raw:
        parts = jid_raw.split("-")
        jf, jb = parts[0].strip(), (parts[1].strip() if len(parts) > 1 else "")
    else:
        jf, jb = jc, ""
    try:    amount = int(amount)
    except: amount = 0
    return dict(name=name, inst=inst, jid_front=jf, jid_back=jb,
                bank=bank, account=account, holder=holder, amount=amount)


def load_upload_wb(work_dir: str) -> openpyxl.Workbook:
    """업로드 양식 로드 — 원본 손상 시 .sb 백업 자동 사용."""
    output = os.path.join(work_dir, "일회성경비지급자_업로드양식_작성.xlsx")
    template = os.path.join(work_dir, "template_일회성경비지급자 업로드양식.xlsx")
    backup_dir = os.path.join(work_dir, "template_일회성경비지급자 업로드양식.xlsx.sb-ce0a4f46-hQO6zg")
    backup = os.path.join(backup_dir, "8C283410.MACTF")

    if os.path.exists(output) and _valid_xlsx(output):
        return openpyxl.load_workbook(output), output

    if os.path.exists(template) and _valid_xlsx(template):
        shutil.copy2(template, output)
    elif os.path.exists(backup):
        with open(backup, "rb") as f: data = f.read()
        with open(output, "wb") as f: f.write(data)
    else:
        raise FileNotFoundError(
            f"업로드 양식 템플릿을 찾을 수 없습니다.\n"
            f"원본: {template}\n백업: {backup}"
        )
    return openpyxl.load_workbook(output), output

def find_end_row(ws) -> int:
    for row in ws.iter_rows(min_col=1, max_col=1):
        if str(row[0].value).strip().upper() == "END":
            return row[0].row
    return ws.max_row + 1


def get_next_seq(wb: openpyxl.Workbook) -> int:
    ws = wb["Sheet1"]
    er = find_end_row(ws)
    if er <= 3: return 1
    try:    return int(ws.cell(er - 1, 1).value) + 1
    except: return max(1, er - 2)


def append_upload_row(wb: openpyxl.Workbook, info: dict, seq: int):
    ws = wb["Sheet1"]
    r  = find_end_row(ws)
    ws.insert_rows(r)
    vals = [
        seq, info["name"], info["inst"],
        info["jid_front"], info["jid_back"],
        f'=IF(H{r}="{DEFAULT_NATIONALITY}","N","Y")',
        "",                    # 여권번호 (내국인 공란)
        DEFAULT_NATIONALITY,
        DEFAULT_INCOME_TYPE,
        DEFAULT_INCOME_DETAIL,
        info["amount"],
        info["account"],
        info["bank"],
        info["holder"],
        0, 0, 0, 0,            # 출장경비 4개
    ]
    for col, val in enumerate(vals, 1):
        ws.cell(r, col).value = val


# ═══════════════════════════════════════════════════════════════════════════
#  페이지 설정 & 디렉토리 선택
# ═══════════════════════════════════════════════════════════════════════════
st.set_page_config(
    page_title="실험참여자비 양식 입력",
    page_icon="🔬",
    layout="centered",
)
st.title("🔬 실험참여자비 양식 자동 입력")

# 세션 초기화: 작업 디렉토리
if "work_dir" not in st.session_state:
    st.session_state.work_dir = os.environ.get(
        "LAB_CHORE_DIR",
        os.path.expanduser("~/Desktop"),
    )

# ── 폴더 선택 바 ──────────────────────────────────────────────────────────
with st.container(border=True):
    col_dir1, col_dir2 = st.columns([5, 1])
    with col_dir1:
        typed = st.text_input(
            "📂 작업 폴더 (템플릿 Excel 파일들이 있는 폴더)",
            value=st.session_state.work_dir,
            label_visibility="collapsed",
            placeholder="폴더 경로를 입력하거나 오른쪽 버튼으로 선택하세요",
            key="dir_input",
        )
        # 타이핑으로 직접 수정 반영
        if typed != st.session_state.work_dir and os.path.isdir(typed):
            st.session_state.work_dir = typed

    with col_dir2:
        if st.button("📂 찾기", use_container_width=True):
            chosen = pick_folder_macos(st.session_state.work_dir)
            if chosen:
                st.session_state.work_dir = chosen
                st.rerun()

WORK_DIR       = st.session_state.work_dir
TEMPLATE_FILE  = os.path.join(WORK_DIR, "실험참여자비 양식(중견).xlsx")

if not os.path.isdir(WORK_DIR):
    st.error("❌ 유효하지 않은 폴더입니다. 올바른 경로를 입력하거나 선택하세요.")
    st.stop()

template_ok = os.path.exists(TEMPLATE_FILE)
if template_ok:
    st.caption(f"✅ 템플릿 확인: `{WORK_DIR}`")
else:
    st.warning(
        f"⚠️ `실험참여자비 양식(중견).xlsx` 템플릿이 이 폴더에 없습니다.  \n"
        f"양식 저장은 불가하지만 업로드 양식 생성은 가능합니다."
    )

# ═══════════════════════════════════════════════════════════════════════════
#  탭 구성
# ═══════════════════════════════════════════════════════════════════════════
tab1, tab2 = st.tabs(["📝  참가자 양식 작성", "📊  업로드 양식 생성"])

# ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
#  TAB 1 — 참가자 양식 작성
# ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
with tab1:
    if not template_ok:
        st.error("템플릿 파일이 없어 양식을 작성할 수 없습니다. 올바른 폴더를 선택하세요.")
        st.stop()

    # ── 참가자 정보 ──────────────────────────────────────────────────────
    st.subheader("👤 참가자 정보")
    c1, c2 = st.columns(2)
    with c1: name  = st.text_input("이름 *", placeholder="예: 홍길동", key="t1_name")
    with c2: inst  = st.text_input("소속 *", value="서울대학교", key="t1_inst")
    c3, c4 = st.columns(2)
    with c3: jid   = st.text_input("주민등록번호 *", placeholder="XXXXXX-XXXXXXX", key="t1_jid")
    with c4: email = st.text_input("이메일 *", placeholder="user@snu.ac.kr", key="t1_email")

    # ── 계좌 정보 ────────────────────────────────────────────────────────
    st.subheader("🏦 계좌 정보")
    c5, c6 = st.columns(2)
    with c5: bank    = st.selectbox("은행명 *", options=BANK_LIST, index=2, key="t1_bank")
    with c6: account = st.text_input("계좌번호 *", placeholder="110-545-811341", key="t1_acc")
    holder = st.text_input("예금주", placeholder="비워두면 이름과 동일", key="t1_holder")

    # ── 활용 정보 ────────────────────────────────────────────────────────
    st.subheader("📋 활용 정보")
    c7, c8 = st.columns([1, 2])
    with c7: amount_str = st.text_input("지급액 * (원)", placeholder="90000", key="t1_amount")
    with c8: date_str   = st.text_input("활용일자", placeholder="2026.03.19~03.20", key="t1_date")

    st.markdown("**⏱ 참여 시간**")
    c9, c10, c11 = st.columns(3)
    with c9:  start_time = st.time_input("시작 시간", value=datetime.time(14, 0), step=1800, key="t1_st")
    with c10: end_time   = st.time_input("종료 시간", value=datetime.time(15, 0), step=1800, key="t1_et")
    with c11:
        duration_h = st.number_input(
            "총 참여 시간 (시간)",
            min_value=0.5, max_value=24.0, value=1.0, step=0.5,
            help="예) 6회×1시간=6",
            key="t1_dur",
        )

    # ── 전자서명 ─────────────────────────────────────────────────────────
    st.subheader("✍️ 전자서명")
    st.caption("캔버스에 서명하면 Excel 파일의 '수령인 서명' 칸에 자동 삽입됩니다.")

    try:
        from streamlit_drawable_canvas import st_canvas
        sig_canvas = st_canvas(
            fill_color="rgba(255,255,255,1)",
            stroke_width=2,
            stroke_color="#111111",
            background_color="#ffffff",
            height=130,
            width=430,
            drawing_mode="freedraw",
            key="t1_canvas",
            display_toolbar=False,
        )
        if st.button("🗑 서명 지우기", key="t1_clr"):
            st.session_state["_sig_v"] = not st.session_state.get("_sig_v", False)
            st.rerun()
        sig_data = sig_canvas.image_data
        has_sig  = (sig_data is not None
                    and sig_data.sum() > 0
                    and not (sig_data[:, :, :3] == 255).all())
        if has_sig: st.success("✅ 서명이 입력되었습니다.")
        else:       st.info("ℹ️ 서명을 그려주세요 (선택사항).")
        CANVAS_OK = True
    except ImportError:
        st.warning("`streamlit-drawable-canvas` 미설치 — `pip install streamlit-drawable-canvas Pillow`")
        sig_data, has_sig, CANVAS_OK = None, False, False

    # ── 저장 버튼 ────────────────────────────────────────────────────────
    st.divider()
    if st.button("✅  양식 저장", type="primary", use_container_width=True, key="t1_save"):

        errors = []
        if not name.strip():    errors.append("이름을 입력하세요.")
        if not inst.strip():    errors.append("소속을 입력하세요.")
        if not jid.strip():     errors.append("주민등록번호를 입력하세요.")
        elif not re.match(r"^\d{6}-\d{7}$", jid.strip()):
            errors.append("주민등록번호 형식: XXXXXX-XXXXXXX")
        if not email.strip():   errors.append("이메일을 입력하세요.")
        if not account.strip(): errors.append("계좌번호를 입력하세요.")
        if not amount_str.strip(): errors.append("지급액을 입력하세요.")
        else:
            try:    amount = int(amount_str.strip().replace(",", ""))
            except: errors.append("지급액은 숫자로 입력하세요.")
        if end_time <= start_time: errors.append("종료 시간이 시작 시간보다 늦어야 합니다.")

        if errors:
            for e in errors: st.error(e)
            st.stop()

        amount       = int(amount_str.strip().replace(",", ""))
        final_holder = holder.strip() or name.strip()
        out_name     = f"실험참여자비 양식_{name.strip()}.xlsx"
        out_path     = os.path.join(WORK_DIR, out_name)

        try:
            shutil.copy2(TEMPLATE_FILE, out_path)
            wb = openpyxl.load_workbook(out_path)
            ws = wb.active

            ws["B16"] = name.strip()
            ws["D16"] = inst.strip()
            ws["E16"] = jid.strip()
            ws["F16"] = email.strip()
            ws["G16"] = bank
            ws["I16"] = account.strip()
            ws["L16"] = final_holder
            ws["D19"] = amount
            if date_str.strip(): ws["C10"] = date_str.strip()
            ws["G10"] = start_time
            ws["I10"] = end_time
            ws["B11"] = int(duration_h) if duration_h == int(duration_h) else duration_h

            if has_sig and sig_data is not None:
                import numpy as np
                sig_pil = Image.fromarray(sig_data.astype("uint8"), "RGBA")
                bg      = Image.new("RGB", sig_pil.size, (255, 255, 255))
                bg.paste(sig_pil, mask=sig_pil.split()[3])
                buf = io.BytesIO()
                bg.save(buf, format="PNG")
                buf.seek(0)
                xl_img         = XLImage(buf)
                xl_img.width   = 160
                xl_img.height  = 55
                xl_img.anchor  = "B17"
                ws.add_image(xl_img)

            wb.save(out_path)
            st.success(f"✅ 저장 완료: **{out_name}**")
            st.info(f"📂 `{out_path}`")
            with st.expander("입력 내용 확인"):
                st.table({"항목": ["이름","소속","주민등록번호","이메일","은행명",
                                   "계좌번호","예금주","지급액","활용일자",
                                   "시작","종료","총시간","서명"],
                          "값":   [name.strip(),inst.strip(),jid.strip(),email.strip(),bank,
                                   account.strip(),final_holder,f"{amount:,}원",
                                   date_str.strip() or "(미입력)",
                                   start_time.strftime("%H:%M"),end_time.strftime("%H:%M"),
                                   f"{duration_h}시간",
                                   "✅ 포함" if has_sig else "—"]})
        except Exception as e:
            import traceback
            st.error(f"저장 실패: {e}")
            st.code(traceback.format_exc())


# ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
#  TAB 2 — 업로드 양식 생성
# ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
with tab2:
    st.subheader("📊 업로드 양식 일괄 생성")
    st.caption(
        "선택된 폴더 내 모든 **실험참여자비 양식_\\*.xlsx** 파일을 읽어 "
        "**일회성경비지급자_업로드양식_작성.xlsx** 한 파일로 통합합니다."
    )

    # 파일 스캔
    found_files = scan_participant_forms(WORK_DIR)
    out_upload  = os.path.join(WORK_DIR, "일회성경비지급자_업로드양식_작성.xlsx")

    if not found_files:
        st.info("ℹ️ 현재 폴더에 `실험참여자비 양식_*.xlsx` 파일이 없습니다.\n\n"
                "탭 1에서 참가자 양식을 먼저 저장하세요.")
    else:
        st.markdown(f"**발견된 파일 {len(found_files)}개:**")
        # 파일 목록 + 이미 처리 여부 표시
        existing_names: set[str] = set()
        if os.path.exists(out_upload) and _valid_xlsx(out_upload):
            try:
                wb_ex = openpyxl.load_workbook(out_upload, data_only=True)
                ws_ex = wb_ex["Sheet1"]
                for row in ws_ex.iter_rows(min_row=3, max_col=2, values_only=True):
                    if row[0] and str(row[0]).upper() != "END" and row[1]:
                        existing_names.add(str(row[1]).strip())
            except Exception:
                pass

        preview_rows = []
        for fp in found_files:
            fname = unicodedata.normalize("NFC", os.path.basename(fp))
            try:
                info = read_participant_info(fp)
                already = "✅ 등록됨" if info["name"] in existing_names else "🔄 신규"
                preview_rows.append({
                    "파일명": fname,
                    "이름":   info["name"],
                    "은행":   info["bank"],
                    "계좌번호": info["account"],
                    "지급액":  f"{info['amount']:,}원",
                    "상태":   already,
                })
            except Exception as e:
                preview_rows.append({"파일명": fname, "이름": "(읽기 오류)",
                                     "은행": "", "계좌번호": "", "지급액": "", "상태": f"❌ {e}"})

        st.dataframe(preview_rows, use_container_width=True, hide_index=True)

        new_count = sum(1 for r in preview_rows if r["상태"] == "🔄 신규")
        st.markdown(f"신규 추가 예정: **{new_count}명** / 이미 등록: **{len(existing_names)}명**")

        col_btn1, col_btn2 = st.columns(2)
        with col_btn1:
            gen_btn = st.button(
                f"📥  업로드 양식 생성 ({new_count}명 추가)",
                type="primary",
                use_container_width=True,
                disabled=(new_count == 0),
                key="t2_gen",
            )
        with col_btn2:
            reset_btn = st.button(
                "🔄  전체 초기화 후 재생성",
                use_container_width=True,
                key="t2_reset",
                help="기존 업로드 양식을 삭제하고 모든 파일로 새로 생성합니다.",
            )

        # ── 초기화 후 재생성 ─────────────────────────────────────────
        if reset_btn:
            if os.path.exists(out_upload):
                os.remove(out_upload)
            st.session_state["_reset_done"] = True
            st.rerun()

        # ── 업로드 양식 생성 ─────────────────────────────────────────
        if gen_btn or st.session_state.get("_reset_done"):
            st.session_state["_reset_done"] = False
            added = skipped = 0
            log_lines = []

            try:
                wb_up, up_path = load_upload_wb(WORK_DIR)
                ws_up = wb_up["Sheet1"]

                # 최신 등록 목록 재로드
                cur_names: set[str] = set()
                for row in ws_up.iter_rows(min_row=3, max_col=2, values_only=True):
                    if row[0] and str(row[0]).upper() != "END" and row[1]:
                        cur_names.add(str(row[1]).strip())

                for fp in found_files:
                    fname = unicodedata.normalize("NFC", os.path.basename(fp))
                    try:
                        info = read_participant_info(fp)
                    except Exception as e:
                        log_lines.append(f"❌ {fname}: {e}")
                        skipped += 1
                        continue
                    if not info["name"]:
                        log_lines.append(f"⚠️ {fname}: 이름 없음, 건너뜀")
                        skipped += 1
                        continue
                    if info["name"] in cur_names:
                        log_lines.append(f"↩️ {info['name']}: 이미 등록됨")
                        skipped += 1
                        continue
                    append_upload_row(wb_up, info, get_next_seq(wb_up))
                    cur_names.add(info["name"])
                    log_lines.append(f"✅ {info['name']} | {info['amount']:,}원 | {info['bank']} {info['account']}")
                    added += 1

                wb_up.save(up_path)

                st.success(f"✅ **완료**: {added}명 추가 / {skipped}건 건너뜀")
                st.info(f"📄 저장: `{up_path}`")

                with st.expander("처리 로그"):
                    for line in log_lines:
                        st.write(line)

                # 최종 테이블 미리보기
                wb_preview = openpyxl.load_workbook(up_path, data_only=True)
                ws_preview = wb_preview["Sheet1"]
                rows_data  = []
                for row in ws_preview.iter_rows(min_row=3, max_col=14, values_only=True):
                    if row[0] is None or str(row[0]).upper() == "END":
                        break
                    rows_data.append({
                        "순번": row[0], "성명": row[1], "소속": row[2],
                        "주민앞": row[3], "주민뒤": row[4],
                        "국적": row[7], "소득구분": row[8],
                        "지급액": row[10], "계좌번호": row[11],
                        "은행": row[12], "예금주": row[13],
                    })
                if rows_data:
                    st.subheader("📋 최종 업로드 양식 미리보기")
                    st.dataframe(rows_data, use_container_width=True, hide_index=True)

            except FileNotFoundError as e:
                st.error(str(e))
            except Exception as e:
                import traceback
                st.error(f"오류 발생: {e}")
                st.code(traceback.format_exc())
